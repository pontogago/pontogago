import re
import openpyxl as px

# Função para extrair informações usando regex
def extrair_informacoes(texto):
    # Padrão regex para encontrar a sequência desejada
    padrao = r"JG STILO MARCENARIA LTDA.*?(\d{6}).*?(\d{2}/\d{2}/\d{4})"
    
    # Encontra todas as correspondências
    correspondencias = re.findall(padrao, texto, re.DOTALL)
    
    # Retorna a primeira correspondência (se houver)
    if correspondencias:
        return correspondencias[0]
    else:
        return None

# Abre o arquivo de texto
with open(r'C:\Users\adcp church\Desktop\gago\python\nomes_arquivos.txt', 'r',  encoding='ISO-8859-1') as arquivo:
    texto = arquivo.read()

# Extrai informações usando a função
informacoes = extrair_informacoes(texto)

# Cria uma planilha usando openpyxl
if informacoes:
    wb = px.Workbook()
    ws = wb.active

    # Insere informações na planilha
    ws['A1'] = "Número"
    ws['B1'] = "Data"
    ws['A2'] = int(informacoes[0])  # Converte o número para inteiro
    ws['B2'] = informacoes[1]  # Mantém a data como string

    # Salva a planilha
    wb.save('informacoes.xlsx')
else:
    print("Nenhuma informação correspondente encontrada no arquivo.")
